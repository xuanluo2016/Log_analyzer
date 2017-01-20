#encoding=utf-8
# by I316639

from __future__ import division
from optparse import OptionParser
from openpyxl import Workbook

import re
import string
import statistics
import prettytable
import operator

def get_args():
    def get_parser():
        usage = """%prog -f filename 

Usage:   JDBC log analysis. Output in jdbc.xlsx and ordered by total running time of each query
Example: python loganalyzer.py -f jdbc.log """
        return OptionParser(usage)
  
    def add_option(parser):
        parser.add_option("-f", "--file", dest="filename",
           help=u"jdbc logging file name", metavar="FILE")        
  
    def get_options(parser):
        options, args = parser.parse_args()
        if not options.filename:
            parser.error('please specify a filename by using parameter -f')
        return options
  
    parser = get_parser()
    add_option(parser)
    return get_options(parser)
  
options      = get_args()
filename     = options.filename
keys         = {}
querycounter = {}
subquerys    = {}
totalcount   = 0
totaltime    = 0
querys       = []
exceptions   = []
exception_id = -1

def split_line(line):
    splitedline = line.split('|')
    return splitedline
    
def get_time(time):
    temp = re.findall("\d",time)
    time = "".join(temp)
    if time.isdigit():
        return (int(time))
    else:
        return exception_id

def save_query(splitedline, totaltime, totalcount):
    if(len(splitedline) < 7):
        exceptions.append(splitedline)
        return totaltime, totalcount
    if("statement" in splitedline[4]):
        query = splitedline[5]        
        time = splitedline[3]
        time = get_time(time)
        if (exception_id == time):
            exceptions.append(splitedline)
            return totaltime, totalcount
        else: 
            totaltime += time
            totalcount += 1
            if query not in querycounter:
                querys.append(query)
                subquerylist = [time] 
                subquerys[query] = subquerylist # store subquerys of each query
                querycounter[query] = 1 # save count of each query
            else:
                subquerylist = subquerys[query]
                subquerylist.append(time) # store running time of subquerys of each query
                querycounter[query] +=1
       
    return totaltime, totalcount

# get index by descending order using binary search
def reverse_bisect_right(array, target, left=0, right=None):
    """Return the index where to insert item x in list a, assuming a is sorted in descending order.

    The return value i is such that all e in a[:i] have element >= target, and all element in
    array[i:] have element < target.  So if target already appears in the list, array.insert(target) will
    insert just after the rightmost target already there.

    Optional args left (default 0) and right (default len(a)) bound the
    slice of a to be searched.

    Essentially, the function returns number of elements in a which are >= than target.
    >>> array = [8, 6, 5, 4, 2]
    >>> reverse_bisect_right(array, 5)
    3
    >>> array[:reverse_bisect_right(array, 5)]
    [8, 6, 5]
    """
    if left < 0:
        raise ValueError('lo must be non-negative')
    if right is None:
        right = len(array)
    while left < right:
        mid = (left+right)//2
        if target > array[mid]: right = mid
        else: left = mid+1
    return left
    
# using binary search algorithm to get index
def get_index_bisearch(data_list, sum_time):
    index = 0
    array = []
    if not data_list: # when data_list is empty
        return index
    for row in data_list:
        if(len(row) < 3):
            index = exception_id
            return exception_id
        else:
            array.append(row[2])
    index = reverse_bisect_right(array,sum_time)
    return index

# deprecated
# get index by total time of each query. using inseart algorighm. replaced by binary search in get_index_bisearch.
def get_index_insertsearch(data_list, sum_time):
    index = 0
    if not data_list: # when data_list is empty
        return index
    for row in data_list:
        if(len(row) < 3):
            index = exception_id
            break
        if(row[2] < sum_time):
            index = data_list.index(row) 
            break
        else:
            index = len(data_list)
    return index
    
def get_index(data_list,sum_time):
    return get_index_bisearch(data_list, sum_time)

# save and sort data
def save_data():
    data_list = []
    for item in subquerys:
        # computing data via mathematic   
        subquerylist = subquerys[item]
        query_count = querycounter[item]#count

        #percent count
        if(totalcount!=0):            
            percent_count = format(query_count/totalcount,'.0%')
            #percent_count = format(round(query_count/totalcount),'.2%')
        else:
            percent_count = format(totalcount,'.0%')
       
        sum_time = sum(subquerylist) # total time

        # percent time 
        if(totaltime != 0):
            percent_time = format(sum_time/totaltime, '.0%')
        else:
            percent_time = format(totaltime, '.0%')
        
        min_time = round(min(subquerylist)) # min
        max_time = round(max(subquerylist))  # max
        mean_time = round(statistics.mean(subquerylist))#average

        std_time = round(statistics.pstdev(subquerylist))# standard deviation

        querystring = item # query string
        
        # store data into workbook
        data = [query_count,percent_count,sum_time,percent_time,min_time,max_time,mean_time,std_time,querystring]
        # sort data
        index = get_index(data_list,sum_time)
        if exception_id == index :
            exceptions.append(data)
        else:
            data_list.insert(index,data)
    return data_list

def create_workbook():
    wb = Workbook();
    ws = wb.active
    ws.title = "JDBC Logging Analysis"
    return wb

def as_text(value):
    if value is None:
        return ""
    return str(value)

def format_workbook(wb):
    ws = wb.active
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column].width = length

def initialize_workbook(wb):
    #Set the column names.  Nine columns by default
    ws = wb.active
    ws.cell(row = 1,column =1).value = "count"
    ws.cell(row = 1,column =2).value = "percent count"
    ws.cell(row = 1,column =3).value = "total time"
    ws.cell(row = 1,column =4).value = "percent time"
    ws.cell(row = 1,column =5).value = "min"
    ws.cell(row = 1,column =6).value = "max"
    ws.cell(row = 1,column =7).value = "average"
    ws.cell(row = 1,column =8).value = "standard deviation"
    ws.cell(row = 1,column =9).value = "query"

# store exception data inito workbook
def save_exceptions(wb):
    ws_exception = wb.create_sheet()
    ws_exception.title = "exceptions"
    ws_exception.cell(row = 1,column =1).value = "jdbc logs with incorrect format!"
    ws_exception.cell(row = 2,column =1).value = "   "

    
    # store exceptions into workbook
    for row in exceptions:
        ws_exception.append(row)

# store active data instead of exception data inito workbook
def save_activedata(wb,data_list):
    ws = wb.active
    
    # store data into workbook
    for row in data_list:
        ws.append(row)  

# save and close workbook
def save_workbook(wb,data_list):
    save_exceptions(wb)
    save_activedata(wb,data_list)
    format_workbook(wb)
    wb.save('jdbc.xlsx')

# this function is to print output into console using prettytable. commented by default. 
def print_prettytable(data_list):
    table = prettytable.PrettyTable()
    table.field_names = ["count","percent count","total time","percent time","min","max","average","standard deviation","query"] 

    # store data into workbook
    for row in data_list:
        table.add_row(row)

    print table   
    #print table.get_string(sortby="total time")

with open(filename) as f:
	for line in f:
            # split each line into different columns via '｜'
            splitedline = split_line(line)

            # Store querystiring, running time and count number of each query into different data sets
            [totaltime, totalcount] = save_query(splitedline,totaltime,totalcount)

# store printed data into data list
data_list = save_data()

# create and initialize workbook
wb = create_workbook()
initialize_workbook(wb)
ws = wb.active
   
# save data into workbook
save_workbook(wb,data_list)

# print output in the console
# print_prettytable(data_list)

print " please refer to jdbc.xlsx for the output"



